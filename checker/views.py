from django.http import JsonResponse
import requests
from .models import SocialMediaURL
from selenium import webdriver
from time import sleep
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import re
df=pd.read_excel("C:/Users/dell/Downloads/social_media_checker/checker/test.xlsx")
urls2=df.iloc[:,0].values.tolist()
urls2=[str(i) for i in urls2]


def is_valid_youtube_url(url):
    # Extract video ID from the URL
    video_id = extract_video_id(url)
    if video_id is None:
        return False

    # Create a YouTube Data API client
    api_service_name = "youtube"
    api_version = "v3"
    api_key = "AIzaSyA5X5Dxyn3QbvEcQCLQrEZTWvDA6wrOLeE"  # Replace with your own API key
    youtube = build(api_service_name, api_version, developerKey=api_key)

    try:
        # Send a request to the API to get video details
        response = youtube.videos().list(
            part="snippet",
            id=video_id
        ).execute()

        # Check if the video exists
        if response.get("items"):
            return True
        else:
            return False

    except HttpError:
        return False

def extract_video_id(url):
    # Extract video ID from various URL formats
    video_id = None
    patterns = [
        r"(?:https?:\/\/)?(?:www\.)?youtu\.?be(?:\.com)?\/(?:watch\?v=|embed\/|v\/|shorts\/|playlist\?list=)?([a-zA-Z0-9_-]{11})",
        r"([a-zA-Z0-9_-]{11})"
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            video_id = match.group(1)
            break
    return video_id

# Example usage



from django.shortcuts import render, redirect
from .forms import SocialMediaURLForm
from .forms import facebookURLForm
from .models import facebookURL

from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.shortcuts import render, redirect

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')  # Replace 'dashboard' with your desired redirect URL after successful login
            else:
                error_message = 'Invalid username or password'
        else:
            error_message = 'Invalid form submission'
    else:
        form = AuthenticationForm()
        error_message = ''
    
    context = {'form': form, 'error_message': error_message}
    return render(request, 'login.html', context)

def signup_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')  # Redirect to the login page after successful signup
        else:
            error_message = 'Invalid form submission'
    else:
        form = UserCreationForm()
        error_message = ''
    
    context = {'form': form, 'error_message': error_message}
    return render(request, 'signup.html', context)


def add_socialmediaurl(request):
    if request.method == 'POST':
        form = SocialMediaURLForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            # Replace 'success' with the URL name for the success page
    else:
        form = SocialMediaURLForm()
    
    context = {'form': form}
    return render(request, 'add_socialmediaurl.html', context)
def add_facebookURL(request):
    if request.method == 'POST':
        form = facebookURLForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            
                    
            # Replace 'success' with the URL name for the success page
    else:
        form = facebookURLForm()
    
    context = {'form': form}
    
    return render(request, 'add_facebookurl.html', context)
from django.shortcuts import render
from .models import SocialMediaURL


def socialmediaurl_list(request):
    urls = SocialMediaURL.objects.all()
    youtube_urls = SocialMediaURL.objects.filter(hisposturl__icontains='youtube.com')
    tiktok_urls = SocialMediaURL.objects.filter(hisposturl__icontains='/www.tiktok.com/')
    twitter_urls = SocialMediaURL.objects.filter(hisposturl__icontains='twitter.com/')
    print(twitter_urls)
    context = {'urls':urls,
                'youtube_urls': youtube_urls,
                'twitter_urls': twitter_urls,
               'tiktoc_urls':tiktok_urls}
    return render(request, 'socialmediaurl_list.html',context)
from django.http import HttpResponse
from datetime import date
def refresh_url(request):
    urls = SocialMediaURL.objects.all()
    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    for i in urls:
            if "twitter.com" in str(i):
                status = SocialMediaURL.objects.get(hisposturl=str(i))
                    
                status.status_date=d1
                status.save()
        
            if "youtube.com/watch?v="  in str(i) or "youtube.com/shorts" in str(i) :
                if is_valid_youtube_url(str(i)):
                    status = SocialMediaURL.objects.get(hisposturl=str(i))
                    status.status = 'ON'
                    status.status_date=d1
                    status.save()
                else:
                    status = SocialMediaURL.objects.get(hisposturl=str(i))
                    status.status = 'OFF'
                    status.status_date=d1
                    status.save()
            if "/www.tiktok.com/" in str(i):
                
            
                max_retries = 3
                retry_delay = 2
                
                for j in range(max_retries):
                    try:
                        response = requests.head(str(i))
            
                        
                        if response.status_code == 200:
                            status = SocialMediaURL.objects.get(hisposturl=str(i))
                            status.status = 'ON'
                            status.status_date=d1
                            status.save()
                        else:
                            status = SocialMediaURL.objects.get(hisposturl=str(i))
                            status.status = 'OFF'
                            status.status_date=d1
                            status.save()
                        break
                    except requests.exceptions.ConnectionError:
                        if j < max_retries - 1:
                            time.sleep(retry_delay)
                            continue
                        status = SocialMediaURL.objects.get(hisposturl=str(i))
                        status.status = 'OFF'
                        status.status_date=d1
                        status.save()
    return HttpResponse("Refresh complete")        
def refresh2(request):
    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    usr="6301708990"
    pwd="Naveen@111"
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--incognito')
    options.add_argument("--start-maximized")
    options.add_argument('--headless')
    options.add_experimental_option( "prefs", {'protocol_handler.excluded_schemes.tel': False})
    driver = webdriver.Chrome("C:/chromedriver.exe",chrome_options=options)
    driver.get('https://www.facebook.com/')
    sleep(1)
    username_box = driver.find_element(By.XPATH,'//*[@id="email"]') 
    username_box.send_keys(usr)
    sleep(1)
    password_box = driver.find_element(By.XPATH,'//*[@id="pass"]')
    password_box.send_keys(pwd)
    login_box = driver.find_element('xpath','/html/body/div[1]/div[1]/div[1]/div/div/div/div[2]/div/div[1]/form/div[2]/button')
    login_box.click()
    sleep(2)
    urls = facebookURL.objects.all()
    
    for i in urls:
        
        if "/www.facebook.com" in str(i) or '/m.facebook.com/ in str(i)':
                driver.get(str(i))
                
                
                try:
                    d=driver.find_element(By.XPATH,'//*[@id="facebook"]/body')
                    print("-------")
                    print(d.text)
                    print("-------")
                    if "This content isn't available at the moment" in d.text:
                        print("OFF")
                        status=facebookURL.objects.get(url=str(i))
                        status.status="OFF"
                        status.status_date=d1
                        status.save()
                    else:
                        print("ON")
                        status=facebookURL.objects.get(url=str(i))
                        status.status="ON"
                        status.status_date=d1
                        status.save()
                except:
                    print("ON")
                    status=facebookURL.objects.get(url=str(i))
                    status.status="ON"
                    status.status_date=d1
                    status.save()   
    

                
            
        

def facebookurl_list(request):
    urls = facebookURL.objects.all()
    context={"urls":urls}
    return render(request, 'facebookurl_list.html', context)
from django.shortcuts import render
from .models import facebookURL, SocialMediaURL

from django.shortcuts import render
from .models import facebookURL, SocialMediaURL

def dashboard(request):
    urls = facebookURL.objects.all()
    urls2 = SocialMediaURL.objects.all()
    profile=0
    groups=0
    pages=0
    for i in urls:
        if 'profile' in str(i):
            profile+=1
        elif 'group' in str(i):
            groups+=1
        elif 'pages' in str(i):
            pages+=1
    total = len(urls) + len(urls2)
    fb_status_values = facebookURL.objects.values_list('status', flat=True)
    youtube_status_values = SocialMediaURL.objects.values_list('status', flat=True)

    combined_status_values = list(fb_status_values) + list(youtube_status_values)

    total_on_count = combined_status_values.count('ON')
    total_off_count = combined_status_values.count('OFF')

    fb_on_count = list(fb_status_values).count('ON')
    fb_off_count = list(fb_status_values).count('OFF')

    youtube_on_count = 0
    youtube_off_count = 0
    tiktoc_on_count = 0
    tiktoc_off_count = 0
    twitter_on_count = 0
    twitter_off_count = 0
    y_status=[]
    t_status=[]
    twitter_status=[]
    for i in urls2:
        if "youtube.com/watch?v=" in str(i):
            y_status.append(i.status)
            if i.status=='ON':
                youtube_on_count+=1
            else:
                youtube_off_count+=1
        if "/www.tiktok.com/" in str(i):
            
            t_status.append(i.status)
            if i.status=='ON':
                tiktoc_on_count+=1
            else:
                tiktoc_off_count+=1
        if "twitter.com/" in str(i):
            
            twitter_status.append(i.status)
            if i.status=='ON':
                twitter_on_count+=1
            else:
                twitter_off_count+=1


    
    context = {
        'total_on_count': total_on_count,
        'total_off_count': total_off_count,
        'fb_total_links':len(urls),
        'youtube_total_links':len(y_status),
        'tiktoc_total_links':len(t_status),
        'twitter_total_links':len(twitter_status),
        'fb_on_count': fb_on_count,
        'fb_off_count': fb_off_count,
        'youtube_on_count': youtube_on_count,
        'youtube_off_count': youtube_off_count,
        'tiktoc_on_count': tiktoc_on_count,
        'tiktoc_off_count': tiktoc_off_count,
        'twitter_on_count': twitter_on_count,
        'twitter_off_count': twitter_off_count,
        'total': total,
        'profile':profile,
        'groups':groups,
        'pages':pages

    }

    return render(request, 'dashboard.html', context)
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse

def update_status(request):
    if request.method == 'POST':
        url = request.POST.get('url')
        status = request.POST.get('status')
        print(url)
        print(status)

        # Perform the necessary logic to update the status in your database
        # Replace the following code with your actual implementation
        try:
            report = get_object_or_404(facebookURL, url=url)
            report.status = status
            report.save()
            return JsonResponse({'success': True})

        except facebookURL.DoesNotExist:
            return JsonResponse({'success': False})

    return JsonResponse({'success': False})
def update_status2(request):
    if request.method == 'POST':
        url = request.POST.get('url')
        new_status = request.POST.get('status')

        # Perform the necessary logic to update the status in the backend
        # Replace the following lines with your actual implementation
        # For example, update the status in the database
        # or update a corresponding model object
        
        # Assuming you have a Report model
        report=get_object_or_404(SocialMediaURL, hisposturl=url)
        report = SocialMediaURL.objects.get(hisposturl=url)
        report.status = new_status
        report.save()

        # Return a JSON response indicating the success of the update
        response = {
            'status': 'success',
            'message': 'Status updated successfully.',
        }
        return JsonResponse(response)

    # Handle cases other than POST requests
    response = {
        'status': 'error',
        'message': 'Invalid request method.',
    }
    return JsonResponse(response)
 
 #
def results_all(request):
    #urls = facebookURL.objects.all()
    urls2 = SocialMediaURL.objects.all()
    youtube_on_count = 0
    youtube_off_count = 0
    tiktoc_on_count = 0
    tiktoc_off_count = 0
    twitter_on_count = 0
    twitter_off_count = 0
    y_status=[]
    t_status=[]
    twitter_status=[]
    for i in urls2:
        if "youtube.com/watch?v=" in str(i):
            y_status.append(i.status)
            if i.status=='ON':
                youtube_on_count+=1
            else:
                youtube_off_count+=1
        if "/www.tiktok.com/" in str(i):
            
            t_status.append(i.status)
            if i.status=='ON':
                tiktoc_on_count+=1
            else:
                tiktoc_off_count+=1
        if "twitter.com/" in str(i):
            
            twitter_status.append(i.status)
            if i.status=='ON':
                twitter_on_count+=1
            else:
                twitter_off_count+=1


    
    context = {
        'urls':urls2,
        'youtube_total_links':len(y_status),
        'tiktoc_total_links':len(t_status),
        'twitter_total_links':len(twitter_status),
        'youtube_on_count': youtube_on_count,
        'youtube_off_count': youtube_off_count,
        'tiktoc_on_count': tiktoc_on_count,
        'tiktoc_off_count': tiktoc_off_count,
        'twitter_on_count': twitter_on_count,
        'twitter_off_count': twitter_off_count,

    }

    return render(request, 'result_all.html', context)
#

def downloadreport_all_links(request):

    urls = SocialMediaURL.objects.all()
    youtube_urls = SocialMediaURL.objects.filter(hisposturl__icontains='youtube.com')
    tiktok_urls = SocialMediaURL.objects.filter(hisposturl__icontains='/www.tiktok.com/')
    twitter_urls = SocialMediaURL.objects.filter(hisposturl__icontains='twitter.com/')
    
    
    context = {'urls':urls,
                'youtube_urls': youtube_urls,
                'twitter_urls': twitter_urls,
               'tiktoc_urls':tiktok_urls}
    return render(request, 'download_list.html',context)


import openpyxl
from django.http import HttpResponse, JsonResponse
from .models import SocialMediaURL

import openpyxl
from django.http import HttpResponse
from .models import SocialMediaURL

def download_excel(request):
    # Get the relevant data based on the selected category
    category = request.GET.get('category')
    
    # Retrieve data for the selected category
    data = None
    if category == 'youtube':
        data = SocialMediaURL.objects.filter(hisposturl__icontains='youtube.com')
    elif category == 'twitter':
        data = SocialMediaURL.objects.filter(hisposturl__icontains='twitter.com')
    elif category == 'tiktok':
        data = SocialMediaURL.objects.filter(hisposturl__icontains='/www.tiktok.com/')
    else:
        data = SocialMediaURL.objects.all()

    # Create an Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the data to the worksheet
    worksheet['A1'] = 'SL.NO.'
    worksheet['B1'] = 'Date'
    worksheet['C1'] = 'Who Post URL'
    worksheet['D1'] = 'Who Screenshot'
    worksheet['E1'] = 'His Post URL'
    worksheet['F1'] = 'His Screenshot'
    worksheet['G1'] = 'Reason for Reporting'
    worksheet['H1'] = 'Specific Cause'
    worksheet['I1'] = 'Violated Law'
    worksheet['J1'] = 'Proposed Action'
    worksheet['K1'] = 'Status Date'
    worksheet['L1'] = 'Status'

    for i, report in enumerate(data, start=2):
        worksheet.cell(row=i, column=1, value=i - 1)
        worksheet.cell(row=i, column=2, value=report.date)
        worksheet.cell(row=i, column=3, value=report.whoposturl)
        worksheet.cell(row=i, column=5, value=report.hisposturl)
        worksheet.cell(row=i, column=7, value=report.reasonforreporting)
        worksheet.cell(row=i, column=8, value=report.specificcause)
        worksheet.cell(row=i, column=9, value=report.violated_law)
        worksheet.cell(row=i, column=10, value=report.proposed_action)
        worksheet.cell(row=i, column=11, value=report.status_date)
        worksheet.cell(row=i, column=12, value=report.status)
        if report.whoscreenshotimage:
            img = openpyxl.drawing.image.Image(report.whoscreenshotimage.path)
            img.width = 80
            img.height = 60
            worksheet.column_dimensions[get_column_letter(4)].width = 20
            worksheet.row_dimensions[i].height = 60
            worksheet.add_image(img, f"D{i}")
        if report.hisscreenshotimage:
            img = openpyxl.drawing.image.Image(report.hisscreenshotimage.path)
            img.width = 80
            img.height = 60
            worksheet.column_dimensions[get_column_letter(4)].width = 20
            worksheet.row_dimensions[i].height = 60
            worksheet.add_image(img, f"F{i}")


    # Create the HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


#
def weekly_checkup_all_list(request):
    urls = SocialMediaURL.objects.all()
    context = {'urls':urls}
    return render(request, 'weekly_checkup_all_list.html',context)
#
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

def download_facebookurl_list(request):
    urls = facebookURL.objects.all()
    context = {"urls": urls}
    return render(request, 'downloadfacebookurl_list.html', context)

def download_excel2(request):
    urls = facebookURL.objects.all()

    # Create a new workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write table headers
    headers = [
        "SL NO.",
        "Date",
        "URL",
        "Screenshot",
        "Reason of Reporting",
        "Specific Cause",
        "Digital Act",
        "Important Person Category",
        "Priority Category",
        "Sub Category",
        "Status Date",
        "Status",
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 15
        worksheet.cell(row=1, column=col_num, value=header).alignment = Alignment(horizontal='center', vertical='center')

    # Write table data
    for row_num, url in enumerate(urls, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=row_num, column=2, value=url.date)
        worksheet.cell(row=row_num, column=3, value=url.url)
        worksheet.cell(row=row_num, column=5, value=url.reason_of_reporting)
        worksheet.cell(row=row_num, column=6, value=url.specific_cause)
        worksheet.cell(row=row_num, column=7, value=url.digital_act)
        worksheet.cell(row=row_num, column=8, value=url.imp_person_category)
        worksheet.cell(row=row_num, column=9, value=url.priority_category)
        worksheet.cell(row=row_num, column=10, value=url.subcategory)
        worksheet.cell(row=row_num, column=11, value=url.status_date)
        worksheet.cell(row=row_num, column=12, value=url.status)

        if url.screenshot_image:
            img = openpyxl.drawing.image.Image(url.screenshot_image.path)
            img.width = 80
            img.height = 60
            worksheet.column_dimensions[get_column_letter(4)].width = 20
            worksheet.row_dimensions[row_num].height = 60
            worksheet.add_image(img, f"D{row_num}")

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=facebook_urls.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def weekly_check_facebookurl_list(request):
    urls = facebookURL.objects.all()
    context = {"urls": urls}
    return render(request, 'weekly_check_facebookurl_list.html', context)

from django.db.models import Count, Q





from django.shortcuts import render
from datetime import date, timedelta
from django.db.models import Count
from .models import facebookURL

def monthly_report(request):
    current_month = date.today().month
    current_year = date.today().year
    prev_month = date.today() - timedelta(days=30)  # Assuming 30 days in a month

    # Get all distinct priority categories from the model
    priority_categories = facebookURL.objects.values_list('priority_category', flat=True).distinct()

    report_data = []
    for category in priority_categories:
        category_data = {
            'priority_category': category,
            'id_count': 0,
            'page_count': 0,
            'group_count': 0,
            'links_count': 0,
            'prev_month_id_count': 0,
            'prev_month_page_count': 0,
            'prev_month_group_count': 0,
            'prev_month_links_count': 0
        }

        # Get current month data for the priority category
        current_month_data = facebookURL.objects.filter(priority_category=category, date__month=current_month, date__year=current_year).values('subcategory').annotate(
            count=Count('subcategory')
        )
        for data in current_month_data:
            subcategory = data['subcategory']
            count = data['count']
            if subcategory == 'ID':
                category_data['id_count'] = count
            elif subcategory == 'Page':
                category_data['page_count'] = count
            elif subcategory == 'Group':
                category_data['group_count'] = count
            elif subcategory == 'Links':
                category_data['links_count'] = count

        # Get previous month data for the priority category
        prev_month_data = facebookURL.objects.filter(priority_category=category, date__month=prev_month.month, date__year=prev_month.year).values('subcategory').annotate(
            count=Count('subcategory')
        )
        for data in prev_month_data:
            subcategory = data['subcategory']
            count = data['count']
            if subcategory == 'ID':
                category_data['prev_month_id_count'] = count
            elif subcategory == 'Page':
                category_data['prev_month_page_count'] = count
            elif subcategory == 'Group':
                category_data['prev_month_group_count'] = count
            elif subcategory == 'Links':
                category_data['prev_month_links_count'] = count

        report_data.append(category_data)

    context = {
        'current_month': current_month,
        'current_year': current_year,
        'report_data': report_data
    }

    return render(request, 'monthly_report.html', context)
